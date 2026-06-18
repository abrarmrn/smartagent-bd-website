# -*- coding: utf-8 -*-
"""
Builds SmartAgentBD_AgencyBrain.xlsx — the Google-Sheets source of truth for the
agency bot. Upload to Google Drive -> Open as Google Sheets -> copy the spreadsheet
id into the n8n workflows. Tab names + header rows are the contract the workflow reads.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "SmartAgentBD_AgencyBrain.xlsx")
FONT = "Calibri"
HEADER_FILL = PatternFill("solid", fgColor="1F6F5C")
HEADER_FONT = Font(name=FONT, size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name=FONT, size=10.5, color="222222")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_COLS = {"long_desc", "answer", "rebuttal", "system_prompt", "mock_catalog", "sample_flow", "description", "headline_features", "key_features"}

SHEETS = {
    "services": {
        "headers": ["service_id", "service_name", "category", "short_desc", "long_desc", "key_features", "best_for", "is_active"],
        "rows": [
            ["svc_fb_bot", "Facebook Messenger AI Bot", "Chatbot", "24/7 Bangla AI that answers every DM and closes COD orders.", "A fully automated Messenger sales agent that replies in fluent Bangla/Banglish within seconds, understands intent, recommends products, captures orders, and never sleeps. Replaces a multi-person moderation team.", "24/7 Bangla replies; Auto order capture; Real-time price & stock; Comment auto-reply", "F-commerce, fashion, food, gadgets", "TRUE"],
            ["svc_multichannel", "Multi-Channel AI (FB + Instagram + WhatsApp)", "Chatbot", "One AI brain across Messenger, Instagram DM and WhatsApp.", "Unifies Facebook, Instagram and WhatsApp into a single AI inbox so no message is missed across channels, with shared memory and consistent replies.", "FB + IG + WhatsApp; Unified inbox; Shared memory; Comment-to-DM", "Multi-channel brands", "TRUE"],
            ["svc_voice_vision", "Voice & Vision AI Add-on", "Add-on", "Understands Bangla voice notes and product photos/screenshots.", "Transcribes Bangla/Banglish voice notes with Whisper and reads product screenshots with a vision model to match items and confirm orders.", "Voice note transcription; Photo product match; Screenshot order confirm", "Fashion, beauty, rich catalogs", "TRUE"],
            ["svc_courier_inventory", "Courier API + Live Inventory Sync", "Automation", "Auto-create courier orders and keep stock in sync.", "Integrates Steadfast/courier APIs to auto-create consignments and syncs live inventory so the bot never sells out-of-stock items.", "Steadfast courier API; Live inventory sync; Auto consignment", "High-volume brands (Advanced)", "TRUE"],
            ["svc_setup_training", "Done-for-You Setup & Training", "Service", "We configure, train and launch the bot for you.", "Full onboarding: we load your catalog, train the AI on your tone, wire courier/inventory, QA, and hand over with a dashboard. You do nothing technical.", "Catalog load; Custom persona; QA; Dashboard handover", "All plans", "TRUE"],
        ],
    },
    "pricing": {
        "headers": ["plan_id", "plan_name", "monthly_price_bdt", "setup_fee_bdt", "message_quota", "channels", "headline_features", "token_tier", "support_sla", "is_popular"],
        "rows": [
            ["basic", "Basic", 2999, 3000, 3500, "Facebook", "1 Facebook Page; 24/7 Bangla AI replies; Standard AI token allocation; Auto order capture; Real-time price & stock", "Standard AI token allocation", "Email support - 24-hour response", "FALSE"],
            ["intermediate", "Intermediate", 9999, 7000, 12000, "Facebook, Instagram, WhatsApp", "FB + IG + WhatsApp; Voice & photo understanding; Comment-to-DM auto-reply; Abandoned-cart recovery; Priority routing; Webhook access", "Expanded AI token utilization", "Priority support - 6-hour response", "TRUE"],
            ["advanced", "Advanced", 24999, 15000, 30000, "Facebook, Instagram, WhatsApp", "Steadfast courier API; Live inventory sync; Custom-trained persona; Unlimited voice & vision; Full webhook & REST API; CRM/ERP integrations; Dedicated account manager", "Maximum AI token utilization", "24/7 priority SLA", "FALSE"],
        ],
    },
    "faqs": {
        "headers": ["faq_id", "question", "answer", "tags", "intent", "priority"],
        "rows": [
            ["faq_lang", "Does the bot reply in pure Bangla?", "Yes - it replies in fluent Bangla, Banglish, and English, matching how your customers actually type. Default tone is warm Banglish.", "language,capability", "capability", 1],
            ["faq_channels", "Can it connect WhatsApp and Instagram?", "Yes. Intermediate and Advanced plans cover Facebook, Instagram DM and WhatsApp from one shared AI brain.", "channels,capability", "capability", 2],
            ["faq_setup", "Is setup included?", "We handle the entire setup for you (done-for-you). There is a transparent one-time setup fee per plan; our team loads your catalog, trains the AI, and launches it.", "setup,onboarding", "onboarding", 3],
            ["faq_roi", "What is the ROI?", "Most pages recover the cost from a handful of orders. From BDT 2,999/month the bot replaces a multi-person moderation team and never misses a DM, so you capture sales 24/7.", "pricing,roi", "pricing", 1],
            ["faq_unknown", "What if the bot does not know an answer?", "It never makes things up. If unsure it says 'ek min, check kore boltesi' and escalates to a human, alerting you instantly.", "capability,safety", "objection", 2],
            ["faq_hidden", "Are there any hidden charges?", "Zero hidden charges. You pay the monthly plan plus a transparent one-time setup fee. Nothing else.", "pricing,trust", "pricing", 2],
        ],
    },
    "objections": {
        "headers": ["objection_id", "objection", "rebuttal", "proof_point"],
        "rows": [
            ["obj_price", "It's too expensive.", "One page admin costs BDT 12,000+/month for 8 hours. SmartAgent BD starts at BDT 2,999/month, works 24/7, never takes leave, and closes orders in under 5 seconds. It pays for itself in a few orders.", "Replaces a 3-5 person moderation team"],
            ["obj_diy", "I can do it myself or just hire an admin.", "Admins sleep, miss DMs at 2am, and quit. The AI never misses a message, handles voice notes & photos, and captures every order automatically - consistently, at a fraction of the cost.", "24/7 zero-miss coverage vs human shifts"],
            ["obj_robotic", "Will it sound robotic?", "No - it chats in natural Banglish with your brand tone and emojis, recommends by photo, and confirms COD like your best salesperson. Try the live demo and judge for yourself.", "Live demo proves natural conversation"],
            ["obj_think", "Let me think about it.", "Totally fair. Let me set up a free 1-day demo with YOUR products so you can watch it close a real order - no commitment. Most owners decide after seeing it work.", "Free tailored demo removes risk"],
        ],
    },
    "demo_personas": {
        "headers": ["persona_id", "keyword", "brand_name", "system_prompt", "mock_catalog", "sample_flow"],
        "rows": [
            ["fashion", "fashion, dress, clothing, panjabi, saree", "Dhaka Threads", "You are Dhaka Threads' AI sales rep. Recommend Panjabi/saree/three-piece by photo, suggest the right size, and confirm COD orders in warm Banglish.", '[{"name":"Premium Panjabi","price":1950,"sizes":"M,L,XL","colors":"White,Navy","stock":20},{"name":"Jamdani Saree","price":3500,"colors":"Red,Green","stock":8}]', "Greet -> ask product/size -> recommend by photo -> confirm color/size -> capture name+phone+address -> confirm COD total + delivery."],
            ["food", "food, restaurant, menu, biriyani, table", "Spice Route", "You are Spice Route restaurant's AI host. Show the menu, take table reservations and delivery orders, and share branch timings in friendly Banglish.", '[{"item":"Kacchi Biriyani","price":320},{"item":"Beef Tehari","price":220},{"item":"Borhani","price":60}]', "Greet -> menu or reservation -> for table: party size+date+time+branch -> capture name+phone -> confirm + maps link."],
            ["resort", "resort, hotel, room, booking, stay", "Sylhet Hills Resort", "You are Sylhet Hills Resort's AI booking assistant. Quote room availability, showcase amenities, and push the direct booking link or capture a lead in Banglish.", '[{"room":"Deluxe Hill View","price_per_night":7500},{"room":"Premium Suite","price_per_night":11000}]', "Greet -> ask dates+guests -> quote rooms -> showcase amenities -> share booking link or capture name+phone+dates."],
            ["support", "support, help, order, return, warranty, track", "TechCare BD", "You are TechCare BD's customer support AI. Handle order tracking, returns and warranty FAQs in calm Banglish; escalate complaints to a human.", '[{"topic":"Order tracking"},{"topic":"7-day returns"},{"topic":"Warranty claim"}]', "Greet -> identify issue -> for tracking ask order id -> give status -> resolve or escalate."],
            ["gym", "gym, fitness, membership, trial, trainer", "IronCore Fitness", "You are IronCore Fitness' AI rep. Explain membership plans, show trainer/slot availability, and book a free trial in energetic Banglish.", '[{"plan":"Monthly","price":2000},{"plan":"Quarterly","price":5000},{"plan":"Yearly","price":15000}]', "Greet -> ask fitness goal -> push free trial -> offer slot -> capture name+phone -> confirm trial + reminder."],
            ["skincare", "skincare, cosmetics, serum, sunscreen, cream", "GlowLab BD", "You are GlowLab BD's skincare advisor AI. Recommend products by skin type and give usage steps in caring Banglish. ALWAYS advise a patch test and escalate any allergy/medical concern to a human expert; never give medical advice.", '[{"name":"COSRX Snail 96 Mucin Essence","price":1350},{"name":"Beauty of Joseon Relief Sun SPF50+","price":1150},{"name":"The Ordinary Niacinamide 10%","price":1250}]', "Greet -> ask skin type/concern -> recommend product -> usage steps -> advise patch test -> capture order or escalate allergy."],
            ["leather", "leather, bag, wallet, crossbody, tote, vanity", "Carindale Leather", "You are Carindale Leather's AI sales rep. Showcase bags/wallets, confirm genuine full-grain leather authenticity, give care tips, and capture orders in polished Banglish.", '[{"name":"Classic Crossbody Bag","price":4500,"colors":"Black,Tan,Maroon"},{"name":"Structured Tote Bag","price":6500},{"name":"Executive Bifold Wallet","price":1800}]', "Greet -> browse bags -> confirm authenticity (genuine full-grain + card) -> care tips -> capture name+phone+address -> confirm COD."],
        ],
    },
    "settings": {
        "headers": ["key", "value", "notes"],
        "rows": [
            ["default_language", "banglish", "Default reply language/tone for the agency bot"],
            ["fallback_message", "Ektu wait korun, ami check kore boltesi", "Used when the bot is unsure"],
            ["escalation_telegram_chat_id", "1380272665", "Telegram chat id for lead alerts"],
            ["business_hours", "Sat-Thu 10am-8pm; Fri closed (bot runs 24/7)", "Human team hours"],
            ["human_handoff_keyword", "agent", "If the user types this, escalate to a human"],
            ["booking_calendar_link", "https://smartagentbd.online/#contact", "Demo booking link the bot shares"],
        ],
    },
    "onboarding_steps": {
        "headers": ["step_no", "plan_scope", "title", "description", "who_does_it", "eta"],
        "rows": [
            [0, "all", "Payment & contract", "Client pays setup + first month and signs a 1-page service agreement + data-processing consent.", "Client", "Day 0"],
            [1, "all", "Welcome & kickoff", "We send a welcome email + onboarding form and book a 30-minute kickoff call.", "SmartAgent BD", "Day 0-1"],
            [2, "all", "Asset access", "Client adds us as a partner in Meta Business Manager with scoped (messaging) asset permissions. We never ask for passwords.", "Both", "Day 1"],
            [3, "all", "Discovery & data", "Client fills the onboarding sheet: product catalog, prices, courier, FAQs, tone, escalation contact.", "Client", "Day 1-2"],
            [4, "advanced", "Build & configure", "We clone the workflow, load the catalog, embed to Pinecone, set persona, and wire courier/inventory.", "SmartAgent BD", "Day 2-4"],
            [5, "advanced", "System User token", "We generate a non-expiring Meta System User token scoped to the client's Page/IG/WhatsApp.", "SmartAgent BD", "Day 3"],
            [6, "all", "Internal QA", "We test in sandbox: Bangla replies, voice, vision, order capture, escalation, courier API, inventory sync.", "SmartAgent BD", "Day 4"],
            [7, "all", "Go-live", "We subscribe the Page to the webhook, set status active, and send a test message from a real account.", "SmartAgent BD", "Day 5"],
            [8, "all", "Client UAT & training", "We walk the client through the dashboard, escalation, takeover, and reading captured orders.", "Both", "Day 5-6"],
            [9, "all", "Handover & support", "We send the handover doc, assign an account manager (Advanced), and start the support SLA clock.", "SmartAgent BD", "Day 6+"],
        ],
    },
    "leads": {
        "headers": ["lead_id", "created_at", "name", "channel", "psid_or_contact", "business_type", "page_link", "interested_plan", "pain_point", "status", "next_action"],
        "rows": [],  # write-target; headers only
    },
}


def style_sheet(ws, headers, rows):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 24
    for r, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(r, c, row[c - 1])
            cell.font = DATA_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=(h in WRAP_COLS))
    for c, h in enumerate(headers, 1):
        longest = len(str(h))
        for row in rows:
            longest = max(longest, min(len(str(row[c - 1])), 60))
        ws.column_dimensions[get_column_letter(c)].width = 48 if h in WRAP_COLS else max(12, min(longest + 3, 40))
    ws.freeze_panes = "A2"
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def main():
    wb = Workbook()
    first = True
    # keep a sensible tab order
    order = ["services", "pricing", "faqs", "objections", "demo_personas", "settings", "onboarding_steps", "leads"]
    for name in order:
        spec = SHEETS[name]
        ws = wb.active if first else wb.create_sheet()
        ws.title = name
        first = False
        style_sheet(ws, spec["headers"], spec["rows"])
    wb.save(OUT)
    print("SAVED:", OUT)
    for ws in wb.worksheets:
        print(f"  tab '{ws.title}': {ws.max_row - 1} data rows x {ws.max_column} cols  | headers: {[c.value for c in ws[1]]}")


if __name__ == "__main__":
    main()
