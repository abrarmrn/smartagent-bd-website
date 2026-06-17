# -*- coding: utf-8 -*-
"""
SmartAgent BD — Multi-Niche Sandbox master workbook generator.
Builds a styled, professional .xlsx with 6 tabs (7-niche demo config + 2 new
product niches + inventory + orders + FAQ/flows). Pure mock data.

Long-text columns (system-prompt snippets, bot responses) use wrap_text with
auto-estimated row height; all other rows stay uniform at 20pt.
"""
import os
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_FILE = os.path.join(OUT_DIR, "SmartAgent_MultiNiche_Sandbox.xlsx")

FONT = "Segoe UI"
DATA_ROW_H = 20.0      # uniform height for normal rows
LINE_H = 15.5          # estimated pixels-per-wrapped-line

PALETTE = {
    "config":    ("483D8B", "E7E5F1"),  # Dark Slate Blue
    "skincare":  ("E78EA9", "FBE6EC"),  # Soft Pastel Rose/Pink
    "leather":   ("9C6B3F", "EFE4D7"),  # Warm Tan / Leather Brown
    "inventory": ("3FA796", "DEF1ED"),  # Soft Mint Green
    "orders":    ("4A90D9", "DDEBF9"),  # Clean Light Blue
    "faqs":      ("7E6BC4", "EBE4F7"),  # Lavender / Purple
}

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def align_for(header: str) -> str:
    h = header.strip().lower()
    if h.endswith("_bdt") or "price" in h or "subtotal" in h or "total" in h or "charge" in h:
        return "right"
    centered = {"product_id", "order_id", "status", "courier_status",
                "stock_qty", "reorder_level", "key/keyword"}
    if h in centered or h.endswith("_id"):
        return "center"
    return "left"


def is_currency(header: str) -> bool:
    h = header.strip().lower()
    return h.endswith("_bdt") or "price" in h


def est_lines(text, width_chars) -> int:
    """Rough estimate of wrapped lines for a given column width (in chars)."""
    if not text or width_chars <= 1:
        return 1
    usable = max(1, width_chars - 1)
    return max(1, math.ceil(len(str(text)) / usable))


def build_sheet(wb, title, scheme_key, headers, rows, first=False,
                wrap_cols=None, wrap_widths=None):
    wrap_cols = set(wrap_cols or [])
    wrap_widths = wrap_widths or {}
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    header_hex, band_hex = PALETTE[scheme_key]
    ws.sheet_properties.tabColor = header_hex
    ws.sheet_view.showGridLines = True

    header_fill = PatternFill("solid", fgColor=header_hex)
    band_fill = PatternFill("solid", fgColor=band_hex)
    header_font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=FONT, size=10.5, color="2B2B2B")

    # ---- column widths (decide first so we can estimate wrap heights) ----
    widths = {}
    for c, head in enumerate(headers, start=1):
        if head in wrap_cols:
            widths[c] = wrap_widths.get(head, 50)
        else:
            longest = len(str(head))
            for row in rows:
                longest = max(longest, len(str(row[c - 1])))
            widths[c] = max(12, min(longest + 3, 62))
        ws.column_dimensions[get_column_letter(c)].width = widths[c]

    # ---- header row ----
    for c, head in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=head)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = BORDER
    ws.row_dimensions[1].height = 26

    # ---- data rows ----
    for r, row in enumerate(rows, start=2):
        banded = (r % 2 == 0)
        max_lines = 1
        for c, head in enumerate(headers, start=1):
            val = row[c - 1]
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = BORDER
            if head in wrap_cols:
                cell.alignment = Alignment(horizontal=align_for(head), vertical="top", wrap_text=True)
                max_lines = max(max_lines, est_lines(val, widths[c]))
            else:
                cell.alignment = Alignment(horizontal=align_for(head), vertical="center", wrap_text=False)
            if banded:
                cell.fill = band_fill
            if is_currency(head) and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
        # uniform 20pt unless this row holds wrapped multi-line text
        ws.row_dimensions[r].height = DATA_ROW_H if max_lines <= 1 else round(max_lines * LINE_H + 4, 1)

    ws.freeze_panes = "A2"
    return ws


# ============================ DATA ============================

config_headers = ["Section", "Key/Keyword", "Persona_Name", "System_Prompt_Snippet", "Status"]
config_rows = [
    ["Demo Niche", "fashion", "Dhaka Threads", "Recommend Panjabi/saree by photo, suggest size, confirm COD in Banglish.", "Active"],
    ["Demo Niche", "food", "Spice Route", "Browse menu, take table reservations & delivery orders, share branch timings.", "Active"],
    ["Demo Niche", "resort", "Sylhet Hills Resort", "Quote room availability, showcase amenities, push direct booking link / capture lead.", "Active"],
    ["Demo Niche", "support", "TechCare BD", "Handle order tracking, returns and warranty FAQs; escalate complaints to human.", "Active"],
    ["Demo Niche", "gym", "IronCore Fitness", "Explain membership plans, show trainer slots, book a free trial.", "Active"],
    ["Demo Niche", "skincare", "GlowLab BD", "Recommend by skin type, give usage steps, ALWAYS flag patch-test & escalate allergies.", "Active"],
    ["Demo Niche", "leather", "Carindale Leather", "Showcase bags/wallets, confirm genuine-leather authenticity, give care tips, capture order.", "Active"],
]

skincare_headers = ["product_id", "brand", "name", "product_type", "skin_type_target",
                    "price_bdt", "key_ingredients", "usage_instruction", "image_filename"]
skincare_rows = [
    ["SKN-001", "COSRX", "Advanced Snail 96 Mucin Power Essence", "Essence", "All / Dry",
     1350, "Snail Secretion Filtrate 96%, Sodium Hyaluronate", "After toner, pat a few drops onto clean face, AM & PM.", "cosrx_snail_mucin.png"],
    ["SKN-002", "Beauty of Joseon", "Relief Sun Rice + Probiotics SPF50+ PA++++", "Sunscreen", "All / Sensitive",
     1150, "Rice Extract, Grain Ferment, Niacinamide", "Apply as final AM step; reapply every 2-3 hours.", "boj_relief_sun.png"],
    ["SKN-003", "The Ordinary", "Niacinamide 10% + Zinc 1%", "Serum", "Oily / Acne-prone",
     1250, "Niacinamide, Zinc PCA", "Apply a few drops AM/PM before moisturizer.", "ordinary_niacinamide.png"],
    ["SKN-004", "CeraVe", "Foaming Facial Cleanser", "Cleanser", "Oily / Combination",
     1450, "Ceramides, Niacinamide, Hyaluronic Acid", "Massage onto wet skin, rinse; AM & PM.", "cerave_foaming_cleanser.png"],
    ["SKN-005", "SKIN1004", "Madagascar Centella Ampoule", "Ampoule", "Sensitive / Irritated",
     1550, "Centella Asiatica Extract 100%", "Apply after toner; pat gently until absorbed.", "skin1004_centella_ampoule.png"],
    ["SKN-006", "Anua", "Heartleaf 77% Soothing Toner", "Toner", "Sensitive / Combination",
     1650, "Houttuynia Cordata Extract 77%, Panthenol", "Soak a cotton pad or pat in by hand, AM/PM.", "anua_heartleaf_toner.png"],
]

leather_headers = ["product_id", "name", "category", "material_type", "price_bdt",
                   "dimensions", "available_colors", "strap_type", "image_filename"]
leather_rows = [
    ["LTR-001", "Classic Crossbody Bag", "Crossbody", "Full-grain Leather", 4500,
     "24 x 18 x 8 cm", "Black, Tan, Maroon", "Adjustable Sling", "leather_crossbody_black.png"],
    ["LTR-002", "Executive Bifold Wallet", "Wallet", "Top-grain Leather", 1800,
     "11 x 9 x 2 cm", "Brown, Black", "N/A", "leather_bifold_wallet.png"],
    ["LTR-003", "Structured Tote Bag", "Tote", "Vegetable-tanned Leather", 6500,
     "35 x 28 x 12 cm", "Tan, Black, Olive", "Dual Top Handle", "leather_tote_black.png"],
    ["LTR-004", "Vintage Messenger Bag", "Messenger", "Full-grain Buffalo Leather", 7200,
     "38 x 28 x 10 cm", "Distressed Brown, Black", "Adjustable Shoulder", "leather_messenger_brown.png"],
    ["LTR-005", "Mini Vanity Shoulder Bag", "Shoulder", "Genuine Cowhide", 3800,
     "20 x 14 x 7 cm", "Blush, Black, Beige", "Chain + Leather Strap", "leather_vanity_blush.png"],
    ["LTR-006", "Slim Card Holder", "Cardholder", "Saffiano Leather", 950,
     "10 x 7 x 1 cm", "Black, Navy, Cherry", "N/A", "leather_cardholder_navy.png"],
]

inv_headers = ["product_id", "product_name", "niche", "size_or_variant", "color",
               "stock_qty", "reorder_level", "warehouse_location"]
inv_rows = [
    ["SKN-001", "COSRX Snail 96 Mucin Essence", "skincare", "100 ml", "N/A", 45, 15, "Dhaka-WH1-A3"],
    ["SKN-002", "BoJ Relief Sun SPF50+", "skincare", "50 ml", "N/A", 32, 12, "Dhaka-WH1-A4"],
    ["SKN-003", "The Ordinary Niacinamide 10%", "skincare", "30 ml", "N/A", 60, 20, "Dhaka-WH1-B1"],
    ["SKN-005", "SKIN1004 Centella Ampoule", "skincare", "55 ml", "N/A", 18, 10, "Dhaka-WH1-B2"],
    ["SKN-006", "Anua Heartleaf Toner", "skincare", "150 ml", "N/A", 27, 12, "Dhaka-WH1-B3"],
    ["LTR-001", "Classic Crossbody Bag", "leather", "Standard", "Black", 12, 5, "Dhaka-WH2-C1"],
    ["LTR-001", "Classic Crossbody Bag", "leather", "Standard", "Tan", 8, 5, "Dhaka-WH2-C1"],
    ["LTR-003", "Structured Tote Bag", "leather", "Standard", "Black", 6, 3, "Dhaka-WH2-C2"],
    ["LTR-002", "Executive Bifold Wallet", "leather", "Standard", "Brown", 25, 8, "Dhaka-WH2-D1"],
    ["LTR-005", "Mini Vanity Shoulder Bag", "leather", "Standard", "Blush", 9, 4, "Dhaka-WH2-C3"],
    ["FSH-001", "Premium Panjabi", "fashion", "L", "Navy", 20, 8, "Dhaka-WH3-E1"],
    ["FSH-001", "Premium Panjabi", "fashion", "M", "White", 14, 8, "Dhaka-WH3-E1"],
]

order_headers = ["order_id", "order_date", "customer_name", "customer_phone", "customer_address",
                 "niche", "items_ordered", "subtotal_bdt", "delivery_charge_bdt", "total_bdt",
                 "courier_status", "courier_tracking_id"]
order_rows = [
    ["ORD-1001", "2026-06-14", "Nusrat Jahan", "017xxxxxxxx", "House 12, Road 5, Dhanmondi, Dhaka",
     "skincare", "COSRX Snail Essence x1, BoJ Sun x1", 2500, 60, 2560, "Delivered", "STDF-7781234"],
    ["ORD-1002", "2026-06-15", "Tanvir Ahmed", "018xxxxxxxx", "Flat 4B, Gulshan-2, Dhaka",
     "leather", "Classic Crossbody (Tan) x1", 4500, 80, 4580, "In Transit", "STDF-7781290"],
    ["ORD-1003", "2026-06-15", "Sadia Islam", "019xxxxxxxx", "Mirpur DOHS, Dhaka",
     "skincare", "Anua Heartleaf Toner x1, Ordinary Niacinamide x1", 2900, 60, 2960, "Processing", "—"],
    ["ORD-1004", "2026-06-16", "Rifat Hossain", "016xxxxxxxx", "Khulshi, Chattogram",
     "leather", "Executive Bifold Wallet x2", 3600, 100, 3700, "Pending", "—"],
    ["ORD-1005", "2026-06-16", "Maliha Rahman", "017xxxxxxxx", "Uttara Sector 7, Dhaka",
     "leather", "Mini Vanity Shoulder Bag (Blush) x1", 3800, 80, 3880, "Confirmed", "STDF-7781355"],
    ["ORD-1006", "2026-06-16", "Arif Chowdhury", "015xxxxxxxx", "Bashundhara R/A, Dhaka",
     "skincare", "CeraVe Cleanser x1, SKIN1004 Ampoule x1", 3000, 60, 3060, "Delivered", "STDF-7781301"],
]

faq_headers = ["niche", "intent", "user_query_sample", "bot_response_template", "human_escalation_trigger"]
faq_rows = [
    ["skincare", "product_recommendation", "amar oily skin, kon serum valo hobe?",
     "Oily skin er jonno Niacinamide 10% serum darun 💧 — oil control kore, pore minimize kore. AM/PM use korben. Aro suggest kori?", "—"],
    ["skincare", "allergy_safety", "amar sensitive skin, allergy ache — eta safe to?",
     "Bujhlam apu/bhaiya 🙏 Sensitive skin e prothome patch test korben (haater kuni te 24 ghonta). Specific kono ingredient e allergy thakle amader skin expert er sathe connect kore dichhi.", "Customer reports allergy / asks medical advice"],
    ["skincare", "usage_instruction", "snail essence ta kibhabe use korbo?",
     "Clean face e, toner er por ar moisturizer er age 💧 koyek drop niye gentle pat korben. AM & PM dono bela use korben.", "—"],
    ["skincare", "order_capture", "ami order korte chai",
     "Joss! 🛍️ Apnar product, naam, phone ar address ta din — COD te confirm kore dichhi.", "—"],
    ["leather", "authenticity", "eta ki original leather naki PU?",
     "Eta 100% genuine full-grain leather ✨ — natural grain ar leather er smell paben, sathe authenticity card o ache. Aro details lagbe?", "Customer disputes authenticity / demands certificate"],
    ["leather", "care_guide", "leather bag er jotno kibhabe nibo?",
     "Sundor proshno 👜 Direct rod ar pani এড়িয়ে cholben, mase ekbar leather conditioner laganen. Bhije gele norom kapor diye muche shukan.", "—"],
    ["leather", "product_browse", "crossbody bag er dam koto?",
     "Amader Classic Crossbody (full-grain) ৳4,500 💼 — Black, Tan ar Maroon ache, adjustable sling strap. Kon color ta dekhabo?", "—"],
    ["general", "human_handoff", "ami akjon manush er sathe kotha bolte chai",
     "Obosshoi 🙏 ami amader ekjon team member ke ekhuni connect kore dichhi, ektu wait korun.", "Explicit human request"],
]

# ============================ BUILD ============================
def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = Workbook()
    build_sheet(wb, "🎯 Sandbox_Config", "config", config_headers, config_rows, first=True,
                wrap_cols=["System_Prompt_Snippet"], wrap_widths={"System_Prompt_Snippet": 52})
    build_sheet(wb, "🧴 Niche_Skincare_Products", "skincare", skincare_headers, skincare_rows)
    build_sheet(wb, "👜 Niche_Leather_Products", "leather", leather_headers, leather_rows)
    build_sheet(wb, "📦 Combined_Inventory", "inventory", inv_headers, inv_rows)
    build_sheet(wb, "🛒 Sandbox_Orders", "orders", order_headers, order_rows)
    build_sheet(wb, "💬 Chatbot_FAQs_&_Flows", "faqs", faq_headers, faq_rows,
                wrap_cols=["user_query_sample", "bot_response_template", "human_escalation_trigger"],
                wrap_widths={"user_query_sample": 30, "bot_response_template": 58, "human_escalation_trigger": 30})
    wb.save(OUT_FILE)
    print("SAVED:", OUT_FILE)
    for ws in wb.worksheets:
        print(f"  - {ws.title}: {ws.max_row-1} data rows x {ws.max_column} cols")


if __name__ == "__main__":
    main()
